from venv import create
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from pkg_resources import working_set
from progress.models import HomeImage,CatalogImage, Logistic,Brend,NewProduct, PriceExcelModel,Product, Category, Certificate, Gallery, Customer, Seller
from django.shortcuts import redirect, render
from django.views.generic import TemplateView, DetailView, ListView
import xlwt
from datetime import datetime

class HomeView(ListView):
    queryset = HomeImage.objects.all()
    template_name = 'index.html'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = NewProduct.objects.filter(type='NEW').order_by('created_at')[:8]
        context['popular_products'] = NewProduct.objects.filter(type='POP').order_by('created_at')[:8]
        return context
  


class AboutView(ListView):
    queryset = Gallery.objects.all()
    template_name = 'about.html'
    context_object_name = 'galeries'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['certificates'] = Certificate.objects.all()[:8]
        context['pricefile'] = PriceExcelModel.objects.all().last()
        return context


class CustomerView(ListView):
    queryset = Customer.objects.all()[:8]
    template_name = 'customers.html'
    context_object_name = 'customers'

class CustomerDetailView(DetailView):
    queryset = Customer.objects.all()
    template_name = 'customer_detail.html'
    context_object_name = 'customer'



class SellerView(ListView):
    queryset = Brend.objects.all()[:8]
    template_name = 'sellers.html'
    context_object_name = 'sellers'

class SellerDetailView(DetailView):
    queryset = Brend.objects.all()
    template_name = 'seller_detail.html'
    context_object_name = 'seller'




class CatalogConstructionView(ListView):
    queryset = Product.objects.all()
    template_name = 'catalog_construction.html'
    context_object_name = 'products1'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.filter(parent='Строительный')
        print(context['categories'])
        context['bacground'] = CatalogImage.objects.get(id=1)
        if self.request.GET.get('seller', None) and self.request.GET.get('category',None):
            context['products'] = Product.objects.filter(seller=self.request.GET.get('seller'),category=self.request.GET.get('category',None),category__parent = 'Строительный')
            context['categories'] = Category.objects.filter(products__seller=self.request.GET.get('seller', None), parent='Строительный')
            context['sellers'] = Seller.objects.all()
            context['cat_id'] = Category.objects.filter(id = self.request.GET.get('category',None)).first().id
        elif self.request.GET.get('seller', None):
            context['sellers'] = Seller.objects.all()
            context['products'] = Product.objects.filter(seller=self.request.GET.get('seller', None), category__parent ='Строительный')
            context['categories'] = Category.objects.filter(products__seller=self.request.GET.get('seller', None), parent='Строительный')
            context['seller_id'] =  Seller.objects.filter(id=self.request.GET.get('seller', None)).first().id
            
        elif self.request.GET.get('category',None):
            context['products'] = Product.objects.filter(category=self.request.GET.get('category',None), category__parent ='Строительный')
            context['categories'] = Category.objects.filter(parent='Строительный')
            context['sellers'] = Seller.objects.all()
            context['cat_id'] = Category.objects.filter(id = self.request.GET.get('category',None)).first().id
        else:
            context['categories'] = Category.objects.filter(parent='Строительный')
            context['sellers'] = Seller.objects.all()
            context['products'] = Product.objects.filter(category__parent='Строительный')
        
        context['new_products'] = NewProduct.objects.filter(type='NEW')
        return context


class CatalogFoodView(ListView):
    queryset = queryset = Product.objects.all()
    template_name = 'catalog_food.html'
    context_object_name = 'products'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.filter(parent='Продукты')
        context['bacground'] = CatalogImage.objects.get(id=2)
        if self.request.GET.get('seller', None) and self.request.GET.get('category',None):
            context['products'] = Product.objects.filter(seller=self.request.GET.get('seller'),category=self.request.GET.get('category',None), category__parent = 'Продукты')
            context['categories'] = Category.objects.filter(products__seller=self.request.GET.get('seller', None), parent='Продукты')
            context['sellers'] = Seller.objects.all()
            context['cat_id'] = Category.objects.filter(id = self.request.GET.get('category',None)).first().id
        
        elif self.request.GET.get('seller', None):
            context['sellers'] = Seller.objects.all()
            context['products'] = Product.objects.filter(seller=self.request.GET.get('seller', None), category__parent = 'Продукты')
            context['categories'] = Category.objects.filter(products__seller=self.request.GET.get('seller', None), parent='Продукты')
            context['seller_id']=  Seller.objects.filter(id=self.request.GET.get('seller', None)).first().id
        
        elif self.request.GET.get('category',None):
            context['products'] = Product.objects.filter(category=self.request.GET.get('category',None),category__parent = 'Продукты')
            context['categories'] = Category.objects.filter(parent='Продукты')
            context['sellers'] = Seller.objects.all()
            context['cat_id'] = Category.objects.filter(id = self.request.GET.get('category',None)).first().id
        else:
            context['categories'] = Category.objects.filter(parent='Продукты')
            context['sellers'] = Seller.objects.all()
            context['products'] = Product.objects.filter(category__parent= 'Продукты')
        
        context['new_products'] = Product.objects.filter(type='NEW')
        return context
    


class ProductTextView(ListView):
    queryset = queryset = Product.objects.all()
    template_name = 'catalog_text.html'
    context_object_name = 'products'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.filter(parent='Текстильные')
        context['bacground'] = CatalogImage.objects.get(id=3)
        if self.request.GET.get('seller', None) and self.request.GET.get('category',None):
            context['products'] = Product.objects.filter(seller=self.request.GET.get('seller'),category=self.request.GET.get('category',None),category__parent ='Текстильные')
            context['categories'] = Category.objects.filter(products__seller=self.request.GET.get('seller', None), parent='Текстильные')
            context['sellers'] = Seller.objects.all()
            context['cat_id'] = Category.objects.filter(id = self.request.GET.get('category',None)).first().id
        elif self.request.GET.get('seller', None):
            context['sellers'] = Seller.objects.all()
            context['products'] = Product.objects.filter(seller=self.request.GET.get('seller', None),category__parent = 'Текстильные' )
            context['categories'] = Category.objects.filter(products__seller=self.request.GET.get('seller', None), parent='Текстильные')
            context['seller_id']=  Seller.objects.filter(id=self.request.GET.get('seller', None)).first().id
            
        elif self.request.GET.get('category',None):
            context['products'] = Product.objects.filter(category=self.request.GET.get('category',None), category__parent = 'Текстильные')
            context['categories'] = Category.objects.filter(parent='Текстильные')
            context['sellers'] = Seller.objects.all()
            context['cat_id'] = Category.objects.filter(id = self.request.GET.get('category',None)).first().id
        else:
            context['categories'] = Category.objects.filter(parent='Текстильные')
            context['sellers'] = Seller.objects.all()
            context['products'] = Product.objects.filter(category__parent = 'Текстильные')

        context['new_products'] = Product.objects.filter(type='NEW')
        return context
    

class HowToByView(ListView):
    queryset = Logistic.objects.all()
    template_name = 'logistics.html'
    context_object_name = 'logistics'
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['logistics_price'] = self.queryset.filter(translations__place_name=self.request.GET.get('logistic',None)).last()
        print(self.request.GET.get('logistic',None))
        context['current_name'] = self.request.GET.get('logistic',None)
        return context

class ContactView(TemplateView):
    template_name ='contact.html'


def export_excelView(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filenmae=Expenses' +str(datetime.date)+ '.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Expenses')
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = ['Наименование', 'Тип_измерения', 'Масса_за_1_шт','Цена']

    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)
    
    font_style = xlwt.XFStyle()
    
    cat_id = request.GET.get('category',None)
    if cat_id:
        rows = Product.objects.filter(category__id=cat_id).values_list(
            'name', 'measurement_type','mass', 'price'
        )
    else:
        rows = Product.objects.all().values_list(
        'name', 'measurement_type','mass', 'price'
        )
    for row in rows:
        row_num += 1
        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)
    wb.save(response)

    return response











def simple_upload(request):
    if request.method == "POST":
        file = request.FILES.get('excel')
        print(file)
        workbook = load_workbook(file)
        sheet = workbook.active
        for i in range(2, sheet.max_row+1):
            name = sheet.cell(row=i, column=1).value
            code = sheet.cell(row=i, column=2).value
            measurement_type = sheet.cell(row=i, column=3).value
            mass = sheet.cell(row=i, column=4).value
            all_mass = sheet.cell(row=i, column=5).value
            price = sheet.cell(row=i, column=6).value
            category = sheet.cell(row=i, column=7).value
            seller = sheet.cell(row=i, column=8).value
            cat_number = sheet.cell(row=i, column=9).value
            model_category = 0
            model_seller =0
            if cat_number and category:
                if cat_number == 1:
                    model_category = Category.objects.create(parent='Строительный', name=category)
                elif cat_number == 2:
                    model_category = Category.objects.create(parent='Продукты', name=category)
                elif cat_number == 3:
                    model_category = Category.objects.create(parent='Текстильные', name=category)
            if seller:
                model_seller = Seller.objects.create(name=seller)
            if model_category and model_seller:
                modelData = Product.objects.create(name=name, code=code, measurement_type=measurement_type, mass=mass, all_mass=all_mass,
                price=price, category=model_category, seller=model_seller)
                modelData.save()

        return redirect('/')
    return render(request, 'upload.html')
